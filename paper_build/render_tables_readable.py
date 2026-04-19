#!/usr/bin/env python3
"""
Render readable PDF/XLSX/CSV tables for:
- Table 1
- Table S.1 to Table S.7
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

import matplotlib.pyplot as plt
import pandas as pd
from matplotlib.backends.backend_pdf import PdfPages


ALPHA_COL = "$\\alpha$"
ZETA_COL = "$\\zeta$"
CITY_GROUPS = ["City 1", "City 2", "City 3"]
CITY_METRICS = ["Accuracy", "Precision", "Recall", "$F_{0.5}$"]
CITY_METRIC_LABELS = ["Accuracy", "Precision", "Recall", r"F$_{0.5}$"]
SIMPLE_METRICS = ["Accuracy", "Precision", "Recall", "$F_{0.5}$"]
SIMPLE_METRIC_LABELS = ["Accuracy", "Precision", "Recall", r"F$_{0.5}$"]

PLAIN_METHOD_MAP_MAIN = {
    "TSLESCAD": "TSLE_SCAD",
    "TSLEMerged": "TSLE_Merged",
    "TSLERandom": "TSLE_Random",
    "TSLEAddition": "TSLE_Addition",
}

PDF_METHOD_MAP_MAIN = {
    "TSLESCAD": r"TSLE$_{SCAD}$",
    "TSLEMerged": r"TSLE$_{Merged}$",
    "TSLERandom": r"TSLE$_{Random}$",
    "TSLEAddition": r"TSLE$_{Addition}$",
}

PLAIN_METHOD_MAP_OTHER = {
    "L0": "TSLE_L0",
    "Lasso": "TSLE_Lasso",
    "MCP": "TSLE_MCP",
}

PDF_METHOD_MAP_OTHER = {
    "L0": r"TSLE$_{L0}$",
    "Lasso": r"TSLE$_{Lasso}$",
    "MCP": r"TSLE$_{MCP}$",
}


@dataclass(frozen=True)
class TableSpec:
    name: str
    title: str
    source_csv: str
    kind: str
    ratio_col: str | None = None
    method_plain_map: dict[str, str] | None = None
    method_pdf_map: dict[str, str] | None = None


TABLE_SPECS = [
    TableSpec("table1", "Table 1", "paper_outputs/tables/table1_wide.csv", "grouped_city", ALPHA_COL, PLAIN_METHOD_MAP_MAIN, PDF_METHOD_MAP_MAIN),
    TableSpec("table_s1", "Table S.1", "paper_outputs/tables/table_s1_wide.csv", "simple"),
    TableSpec("table_s2", "Table S.2", "paper_outputs/tables/table_s2_wide.csv", "grouped_city", ALPHA_COL, PLAIN_METHOD_MAP_OTHER, PDF_METHOD_MAP_OTHER),
    TableSpec("table_s3", "Table S.3", "paper_outputs/tables/table_s3_wide.csv", "grouped_city", ALPHA_COL, PLAIN_METHOD_MAP_OTHER, PDF_METHOD_MAP_OTHER),
    TableSpec("table_s4", "Table S.4", "paper_outputs/tables/table_s4_wide.csv", "grouped_city", ALPHA_COL, PLAIN_METHOD_MAP_OTHER, PDF_METHOD_MAP_OTHER),
    TableSpec("table_s5", "Table S.5", "paper_outputs/tables/table_s5_wide.csv", "grouped_city", ALPHA_COL, PLAIN_METHOD_MAP_OTHER, PDF_METHOD_MAP_OTHER),
    TableSpec("table_s6", "Table S.6", "paper_outputs/tables/table_s6_wide.csv", "grouped_city", ALPHA_COL, PLAIN_METHOD_MAP_MAIN, PDF_METHOD_MAP_MAIN),
    TableSpec("table_s7", "Table S.7", "paper_outputs/tables/table_s7_wide.csv", "grouped_metric", ZETA_COL, PLAIN_METHOD_MAP_MAIN, PDF_METHOD_MAP_MAIN),
]


def clean_cell(text: object) -> str:
    if pd.isna(text):
        return "--"
    return str(text).replace(r"\%", "%").strip()


def split_mean_std(text: object) -> tuple[str, str]:
    cleaned = clean_cell(text)
    if cleaned == "--":
        return "--", ""
    if " (" in cleaned and cleaned.endswith(")"):
        mean, std = cleaned.split(" (", 1)
        return mean, f"({std}"
    return cleaned, ""


def multiline_cell(text: object) -> str:
    mean, std = split_mean_std(text)
    if not std:
        return mean
    return f"{mean}\n{std}"


def ratio_label_for_csv(col_name: str | None) -> str:
    if col_name == ZETA_COL:
        return "zeta"
    return "alpha"


def ratio_label_for_pdf(col_name: str | None) -> str:
    if col_name == ZETA_COL:
        return r"$\zeta$"
    return r"$\alpha$"


def build_grouped_city_readable(df: pd.DataFrame, ratio_col: str, method_plain_map: dict[str, str]) -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    for method in df["Method"].drop_duplicates():
        sub = df[df["Method"] == method].reset_index(drop=True)
        for idx, (_, record) in enumerate(sub.iterrows()):
            mean_row: dict[str, str] = {"Method": method_plain_map.get(method, method), ratio_label_for_csv(ratio_col): f"{float(record[ratio_col]):.2f}"}
            std_row: dict[str, str] = {"Method": "", ratio_label_for_csv(ratio_col): ""}
            if idx > 0:
                mean_row["Method"] = ""
            for city in CITY_GROUPS:
                for metric in CITY_METRICS:
                    mean, std = split_mean_std(record[f"{city} | {metric}"])
                    mean_row[f"{city} {metric}"] = mean
                    std_row[f"{city} {metric}"] = std
            rows.extend([mean_row, std_row])
    return pd.DataFrame(rows)


def build_grouped_metric_readable(df: pd.DataFrame, ratio_col: str, method_plain_map: dict[str, str]) -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    for method in df["Method"].drop_duplicates():
        sub = df[df["Method"] == method].reset_index(drop=True)
        for idx, (_, record) in enumerate(sub.iterrows()):
            mean_row = {"Method": method_plain_map.get(method, method), ratio_label_for_csv(ratio_col): f"{float(record[ratio_col]):.2f}"}
            std_row = {"Method": "", ratio_label_for_csv(ratio_col): ""}
            if idx > 0:
                mean_row["Method"] = ""
            for metric in SIMPLE_METRICS:
                mean, std = split_mean_std(record[metric])
                mean_row[metric] = mean
                std_row[metric] = std
            rows.extend([mean_row, std_row])
    return pd.DataFrame(rows)


def build_simple_readable(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "City" in out.columns:
        out["City"] = out["City"].astype(str)
    return out


def _style_axis(ax: plt.Axes, title: str) -> None:
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis("off")
    ax.text(0.5, 0.975, title, ha="center", va="center", fontsize=14, fontweight="bold")


def render_grouped_city_pdf(df: pd.DataFrame, title: str, ratio_col: str, method_pdf_map: dict[str, str]) -> plt.Figure:
    plt.rcParams["font.family"] = "DejaVu Serif"

    rows = []
    for _, record in df.iterrows():
        rows.append(
            {
                "method": str(record["Method"]),
                "ratio": f"{float(record[ratio_col]):.2f}",
                "values": {
                    city: [multiline_cell(record[f"{city} | {metric}"]) for metric in CITY_METRICS]
                    for city in CITY_GROUPS
                },
            }
        )

    fig_height = max(8.0, 3.8 + 0.42 * len(rows))
    fig, ax = plt.subplots(figsize=(16.5, fig_height))
    _style_axis(ax, title)

    left = 0.02
    right = 0.98
    top = 0.94
    bottom = 0.04
    total_width = right - left

    method_w = 0.12 * total_width
    ratio_w = 0.05 * total_width
    metric_w = (total_width - method_w - ratio_w) / 12.0

    header1_h = 0.055
    header2_h = 0.045
    data_h = (top - bottom - header1_h - header2_h) / len(rows)

    x_edges = [left, left + method_w, left + method_w + ratio_w]
    for _ in range(12):
        x_edges.append(x_edges[-1] + metric_w)

    y_header1 = top - header1_h / 2
    y_header2 = top - header1_h - header2_h / 2

    ax.plot([left, right], [top, top], color="black", lw=1.2)
    ax.plot([left, right], [bottom, bottom], color="black", lw=1.2)
    ax.plot([left, right], [top - header1_h - header2_h, top - header1_h - header2_h], color="black", lw=0.9)

    for group_idx in range(3):
        start = x_edges[2 + group_idx * 4]
        end = x_edges[2 + (group_idx + 1) * 4]
        ax.plot([start + 0.004, end - 0.004], [top - header1_h, top - header1_h], color="black", lw=0.8)

    ax.text((x_edges[0] + x_edges[1]) / 2, y_header1, "Method", ha="center", va="center", fontsize=12)
    ax.text((x_edges[1] + x_edges[2]) / 2, y_header1, ratio_label_for_pdf(ratio_col), ha="center", va="center", fontsize=12)
    for group_idx, city in enumerate(CITY_GROUPS):
        start = x_edges[2 + group_idx * 4]
        end = x_edges[2 + (group_idx + 1) * 4]
        ax.text((start + end) / 2, y_header1, city, ha="center", va="center", fontsize=13)

    for group_idx in range(3):
        for metric_idx, metric in enumerate(CITY_METRIC_LABELS):
            start = x_edges[2 + group_idx * 4 + metric_idx]
            end = x_edges[3 + group_idx * 4 + metric_idx]
            ax.text((start + end) / 2, y_header2, metric, ha="center", va="center", fontsize=10)

    current_top = top - header1_h - header2_h
    group_start = 0
    while group_start < len(rows):
        method = rows[group_start]["method"]
        group_end = group_start
        while group_end < len(rows) and rows[group_end]["method"] == method:
            group_end += 1
        block_rows = rows[group_start:group_end]
        block_center_y = current_top - data_h * len(block_rows) / 2
        ax.text(
            (x_edges[0] + x_edges[1]) / 2,
            block_center_y,
            method_pdf_map.get(method, method),
            ha="center",
            va="center",
            fontsize=12,
            fontweight="bold",
        )
        for row_offset, row in enumerate(block_rows):
            y_center = current_top - data_h * (row_offset + 0.5)
            ax.text((x_edges[1] + x_edges[2]) / 2, y_center, row["ratio"], ha="center", va="center", fontsize=11, fontweight="bold")
            for group_idx, city in enumerate(CITY_GROUPS):
                for metric_idx, value in enumerate(row["values"][city]):
                    start = x_edges[2 + group_idx * 4 + metric_idx]
                    end = x_edges[3 + group_idx * 4 + metric_idx]
                    ax.text((start + end) / 2, y_center, value, ha="center", va="center", fontsize=9)
        current_top -= data_h * len(block_rows)
        ax.plot([left, right], [current_top, current_top], color="black", lw=0.8)
        group_start = group_end

    return fig


def render_grouped_metric_pdf(df: pd.DataFrame, title: str, ratio_col: str, method_pdf_map: dict[str, str]) -> plt.Figure:
    plt.rcParams["font.family"] = "DejaVu Serif"

    rows = []
    for _, record in df.iterrows():
        rows.append(
            {
                "method": str(record["Method"]),
                "ratio": f"{float(record[ratio_col]):.2f}",
                "values": [multiline_cell(record[col]) for col in SIMPLE_METRICS],
            }
        )

    fig_height = max(6.0, 3.2 + 0.45 * len(rows))
    fig, ax = plt.subplots(figsize=(9.5, fig_height))
    _style_axis(ax, title)

    left = 0.04
    right = 0.96
    top = 0.92
    bottom = 0.06
    total_width = right - left

    method_w = 0.2 * total_width
    ratio_w = 0.08 * total_width
    metric_w = (total_width - method_w - ratio_w) / len(SIMPLE_METRICS)
    header_h = 0.06
    data_h = (top - bottom - header_h) / len(rows)

    x_edges = [left, left + method_w, left + method_w + ratio_w]
    for _ in SIMPLE_METRICS:
        x_edges.append(x_edges[-1] + metric_w)

    ax.plot([left, right], [top, top], color="black", lw=1.2)
    ax.plot([left, right], [bottom, bottom], color="black", lw=1.2)
    ax.plot([left, right], [top - header_h, top - header_h], color="black", lw=0.9)

    y_header = top - header_h / 2
    ax.text((x_edges[0] + x_edges[1]) / 2, y_header, "Method", ha="center", va="center", fontsize=12)
    ax.text((x_edges[1] + x_edges[2]) / 2, y_header, ratio_label_for_pdf(ratio_col), ha="center", va="center", fontsize=12)
    for idx, metric in enumerate(SIMPLE_METRIC_LABELS):
        ax.text((x_edges[2 + idx] + x_edges[3 + idx]) / 2, y_header, metric, ha="center", va="center", fontsize=10)

    current_top = top - header_h
    group_start = 0
    while group_start < len(rows):
        method = rows[group_start]["method"]
        group_end = group_start
        while group_end < len(rows) and rows[group_end]["method"] == method:
            group_end += 1
        block_rows = rows[group_start:group_end]
        block_center_y = current_top - data_h * len(block_rows) / 2
        ax.text((x_edges[0] + x_edges[1]) / 2, block_center_y, method_pdf_map.get(method, method), ha="center", va="center", fontsize=12, fontweight="bold")
        for row_offset, row in enumerate(block_rows):
            y_center = current_top - data_h * (row_offset + 0.5)
            ax.text((x_edges[1] + x_edges[2]) / 2, y_center, row["ratio"], ha="center", va="center", fontsize=11, fontweight="bold")
            for idx, value in enumerate(row["values"]):
                ax.text((x_edges[2 + idx] + x_edges[3 + idx]) / 2, y_center, value, ha="center", va="center", fontsize=9)
        current_top -= data_h * len(block_rows)
        ax.plot([left, right], [current_top, current_top], color="black", lw=0.8)
        group_start = group_end

    return fig


def render_simple_pdf(df: pd.DataFrame, title: str) -> plt.Figure:
    plt.rcParams["font.family"] = "DejaVu Serif"
    fig_height = max(3.0, 1.8 + 0.35 * (len(df) + 1))
    fig, ax = plt.subplots(figsize=(7.2, fig_height))
    _style_axis(ax, title)
    table = ax.table(
        cellText=df.values,
        colLabels=df.columns.tolist(),
        cellLoc="center",
        colLoc="center",
        loc="center",
        bbox=[0.08, 0.12, 0.84, 0.72],
    )
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    for (row, col), cell in table.get_celld().items():
        cell.set_edgecolor("black")
        cell.set_linewidth(0.8 if row == 0 else 0.5)
        if row == 0:
            cell.set_text_props(fontweight="bold")
    return fig


def apply_excel_formatting(path: Path) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font
    except Exception:
        return

    wb = load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                if cell.row == 1:
                    cell.font = Font(bold=True)
        for column_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            width = min(max(max_len + 2, 10), 22)
            ws.column_dimensions[column_cells[0].column_letter].width = width
        if ws.max_row <= 8:
            for row_idx in range(1, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 22
        else:
            for row_idx in range(1, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 28
    wb.save(path)


def render_from_dataframe(
    spec: TableSpec,
    df: pd.DataFrame,
    output_dir: Path,
    workbook: pd.ExcelWriter,
    pdf_pages: PdfPages,
) -> None:
    if spec.kind == "grouped_city":
        readable = build_grouped_city_readable(df, spec.ratio_col or ALPHA_COL, spec.method_plain_map or {})
        fig = render_grouped_city_pdf(df, spec.title, spec.ratio_col or ALPHA_COL, spec.method_pdf_map or {})
    elif spec.kind == "grouped_metric":
        readable = build_grouped_metric_readable(df, spec.ratio_col or ZETA_COL, spec.method_plain_map or {})
        fig = render_grouped_metric_pdf(df, spec.title, spec.ratio_col or ZETA_COL, spec.method_pdf_map or {})
    else:
        readable = build_simple_readable(df)
        fig = render_simple_pdf(readable, spec.title)

    csv_path = output_dir / f"{spec.name}_readable.csv"
    xlsx_path = output_dir / f"{spec.name}_readable.xlsx"
    pdf_path = output_dir / f"{spec.name}_readable.pdf"

    readable.to_csv(csv_path, index=False)
    readable.to_excel(xlsx_path, index=False)
    apply_excel_formatting(xlsx_path)
    readable.to_excel(workbook, sheet_name=spec.name[:31], index=False)
    fig.savefig(pdf_path, format="pdf", bbox_inches="tight")
    pdf_pages.savefig(fig, bbox_inches="tight")
    plt.close(fig)


def render_for_spec(spec: TableSpec, output_dir: Path, workbook: pd.ExcelWriter, pdf_pages: PdfPages) -> None:
    df = pd.read_csv(spec.source_csv)
    render_from_dataframe(spec, df, output_dir, workbook, pdf_pages)


def export_tables_from_frames(frames: dict[str, pd.DataFrame], output_dir: Path | str = "paper_outputs/tables/readable") -> Path:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    combined_xlsx = output_dir / "all_readable_tables.xlsx"
    combined_pdf = output_dir / "all_readable_tables.pdf"

    with pd.ExcelWriter(combined_xlsx, engine="openpyxl") as workbook, PdfPages(combined_pdf) as pdf_pages:
        for spec in TABLE_SPECS:
            df = frames.get(spec.name)
            if df is None:
                raise KeyError(f"Missing source dataframe for readable export: {spec.name}")
            render_from_dataframe(spec, df, output_dir, workbook, pdf_pages)

    apply_excel_formatting(combined_xlsx)
    return output_dir


def main() -> None:
    parser = argparse.ArgumentParser(description="Export human-readable CSV/XLSX/PDF tables.")
    parser.add_argument("--output-dir", default="paper_outputs/tables/readable", help="Directory for readable exports.")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    combined_xlsx = output_dir / "all_readable_tables.xlsx"
    combined_pdf = output_dir / "all_readable_tables.pdf"

    with pd.ExcelWriter(combined_xlsx, engine="openpyxl") as workbook, PdfPages(combined_pdf) as pdf_pages:
        for spec in TABLE_SPECS:
            render_for_spec(spec, output_dir, workbook, pdf_pages)

    apply_excel_formatting(combined_xlsx)
    print(output_dir)


if __name__ == "__main__":
    main()
